import datetime
import os
import pathlib
import re
import unittest
from unittest.mock import patch

import excel_shopping_tracker
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

categories = ["baby", "regular groceries", "game", "car related", "taxi"]
temp_file_path = os.path.join('C:/Users', os.environ['USERPROFILE'], 'AppData/Local/Temp/test_Finances.xlsx')
file_path = str(pathlib.Path().absolute()) + '/tests/test1.xlsx'
second_file_path = str(pathlib.Path().absolute()) + '/tests/test2.xlsx'

class TestApp(unittest.TestCase):

    maxDiff = None

    def test_validate_date(self):
        self.assertEqual(excel_shopping_tracker.validate_date("12/12/2020"), True)
        self.assertEqual(excel_shopping_tracker.validate_date("2021/06/22"), False)
        self.assertEqual(excel_shopping_tracker.validate_date("test"), False)

    def test_validate_price(self):
        self.assertEqual(excel_shopping_tracker.validate_price("29.99"), True)
        self.assertEqual(excel_shopping_tracker.validate_price("20"), True)
        self.assertEqual(excel_shopping_tracker.validate_price("yo"), False)
        self.assertEqual(excel_shopping_tracker.validate_price("123.23.34"), False)
    
    def test_ask_question(self):
        # First not allowed input is given, then yes
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["test", "yes"]
            result = excel_shopping_tracker.ask_question("Would you like to add any rows to Excel?")
            self.assertEqual(result, True)
        
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["yes"]
            result = excel_shopping_tracker.ask_question("Would you like to add any rows to Excel?")
            self.assertEqual(result, True)
        
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["no"]
            result = excel_shopping_tracker.ask_question("Would you like to add any rows to Excel?")
            self.assertEqual(result, False)
    
    def test_collect_user_input(self):
        # One row
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["10/12/2020", "123", "2", "Zelda rules", "no"]
            result = excel_shopping_tracker.collect_user_input(categories)
            self.assertEqual(result, [{'date': '10/12/2020', 'amount': 123, 'category': 'regular groceries', 'description': 'Zelda rules'}])
        
        # Wrong formats first
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["10/13/2020", "2021/12/03", "10/12/2019", "one hundred", "100", "6", "4", "2 packages of diapers", "no"]
            result = excel_shopping_tracker.collect_user_input(categories)
            self.assertEqual(result, [{'date': '10/12/2019', 'amount': 100, 'category': 'car related', 'description': '2 packages of diapers'}])
        
        # Multiple rows
        with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = (
                    "10/01/2021", "100.99", "3", "fuel", "yes",
                    "12/01/2021", "23.88", "2", "Mario", "yes",
                    "14/01/2021", "455", "1", "Auchan", "no"
                )
            result = excel_shopping_tracker.collect_user_input(categories)
            self.assertEqual(
                    result,
                    [
                        {'date': '10/01/2021', 'amount': 100.99, 'category': 'game', 'description': 'fuel'},
                        {'date': '12/01/2021', 'amount': 23.88, 'category': 'regular groceries', 'description': 'Mario'}, 
                        {'date': '14/01/2021', 'amount': 455, 'category': 'baby', 'description': 'Auchan'}
                    ]
                )

    def test_create_new_excel(self):       
        excel_shopping_tracker.create_new_excel(temp_file_path)
        assert os.path.isfile(temp_file_path)
        os.remove(temp_file_path)
        
    def setUp(self):
        excel_shopping_tracker.create_new_excel(file_path)
    
    def test_save_new_rows_to_excel(self):
        
        rows = [
            {'date': '01/01/2021', 'amount': 12, 'category': 'regular groceries', 'description': 'test test one'},
            {'date': '01/03/2021', 'amount': 123, 'category': 'regular groceries', 'description': 'test test 2'},
            {'date': '05/01/2021', 'amount': 12, 'category': 'taxi', 'description': 'test test three'},
        ]
        excel_shopping_tracker.save_new_rows_to_excel(rows, file_path)

        # Now verify if the rows were saved
        workbook = load_workbook(file_path)
        sheet = workbook.active
        all_rows = []
        for row in sheet.iter_rows(min_row=2,min_col=1, max_col=4, values_only=True):
            all_rows.append(row)
        self.assertEqual(
                    all_rows,
                    [
                        (datetime.datetime(2021, 1, 1, 0, 0), 12, 'regular groceries', 'test test one'),
                        (datetime.datetime(2021, 3, 1, 0, 0), 123, 'regular groceries', 'test test 2'), 
                        (datetime.datetime(2021, 1, 5, 0, 0), 12, 'taxi', 'test test three')
                    ]
                )
    
    def tearDown(self):
        os.remove(file_path)

    def test_get_stats_options(self):
        options = excel_shopping_tracker.get_stats_options(second_file_path)
        print(options)
        self.assertEqual(options, ['09/20', '10/20', '11/20', '12/20', '03/21', '10/21'])

    def test_get_user_request(self):
         options = ['08/19', '09/20', '10/20', '11/20', '12/20', '03/21', '10/21']
         
         # user does not want statistics
         with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["no"]
            result = excel_shopping_tracker.get_user_request(options)
            self.assertEqual(result, [])
        
        # incorrect input given first
         with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["yup", "yes", "March 20 April 20", "08/19 09/20 10/20", "no"]
            result = excel_shopping_tracker.get_user_request(options)
            self.assertEqual(result, ['08/19', '09/20', '10/20'])
        
        # all good first time
         with patch('builtins.input') as mocked_input:
            mocked_input.side_effect = ["yes", "08/19 09/20 10/20 11/20 12/20", "no"]
            result = excel_shopping_tracker.get_user_request(options)
            self.assertEqual(result, ['08/19', '09/20', '10/20', '11/20', '12/20'])
    
    def test_get_stats_data(self):
        result = excel_shopping_tracker.get_stats_data(
            categories, ['09/20', '10/20', '11/20', '12/20'], second_file_path
        )
        
        self.assertEqual(
            result,
            [
                {'month': '09/20', 'total': 23, 'top_5': [(datetime.datetime(2020, 9, 13, 0, 0), 23, 'game', 'test 123')],
                'categories': {'baby': 0, 'regular groceries': 0, 'game': 23, 'car related': 0, 'taxi': 0}},

                {'month': '10/20', 'total': 120, 'top_5': [(datetime.datetime(2020, 10, 16, 0, 0), 120, 'taxi', 'test 126')],
                'categories': {'baby': 0, 'regular groceries': 0, 'game': 0, 'car related': 0, 'taxi': 120}},

                {'month': '11/20', 'total': 232, 'top_5': [(datetime.datetime(2020, 11, 18, 0, 0), 232, 'baby', 'test 128')],
                'categories': {'baby': 232, 'regular groceries': 0, 'game': 0, 'car related': 0, 'taxi': 0}},

                {'month': '12/20', 'total': 525,
                'top_5': [(datetime.datetime(2020, 12, 12, 0, 0), 345, 'taxi', 'sdfsdfsdfsdf'), (datetime.datetime(2020, 12, 17, 0, 0), 123, 'baby', 'test 127'),
                (datetime.datetime(2020, 12, 14, 0, 0), 34, 'groceries', 'test 124'), (datetime.datetime(2020, 12, 15, 0, 0), 23, 'groceries', 'test 125')],
                'categories': {'baby': 123, 'regular groceries': 0, 'game': 0, 'car related': 0, 'taxi': 345}}
            ]
        )
    
    def test_prepare_message(self):
        stats_data = [
            {'month': '09/20', 'total': 23, 'top_5': [(datetime.datetime(2020, 9, 13, 0, 0), 23, 'game', 'test 123')],
                'categories': {'baby': 0, 'regular groceries': 0, 'game': 23, 'car related': 0, 'taxi': 0}},
        ]

        expected_message = """
Here are statistics for 09/20:

The total spent: 23

The highest transactions of the month: 
1: 13/09/20, amount: 23, category: game, details: test 123

Totals of each category:
baby: 0
regular groceries: 0
game: 23
car related: 0
taxi: 0
"""
            

        result = excel_shopping_tracker.prepare_message(stats_data)

        self.assertEqual(result, expected_message)

    def test_plot(self):
        stats_data1 = [
                {'month': '09/20', 'total': 23, 'top_5': [(datetime.datetime(2020, 9, 13, 0, 0), 23, 'game', 'test 123')],
                'categories': {'baby': 0, 'regular groceries': 0, 'game': 23, 'car related': 0, 'taxi': 0}},

                {'month': '10/20', 'total': 120, 'top_5': [(datetime.datetime(2020, 10, 16, 0, 0), 120, 'taxi', 'test 126')],
                'categories': {'baby': 0, 'regular groceries': 0, 'game': 0, 'car related': 0, 'taxi': 120}},
        ]
    
        with patch("excel_shopping_tracker.plt.show") as show_patch:
            excel_shopping_tracker.plot(categories, stats_data1)
            assert show_patch.called

if __name__ == '__main__':
    unittest.main()
