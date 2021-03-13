import os
import re
from datetime import datetime

import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

file_path = os.path.join(os.environ['USERPROFILE'], 'Desktop/Finances.xlsx')
categories = ["baby", "regular groceries", "game", "car related", "taxi"]

def validate_date(user_input: str) -> bool:
    '''Validates whether user inputs date in correct form'''

    try:
        datetime.strptime(user_input, "%d/%m/%Y")
        return True
    except ValueError:
        print("Incorrect date format, should be dd/mm/yyyy")
        return False

def validate_price(user_input: str) -> bool:
    match = re.search(r"^\d+(\.\d{2})?$", user_input)
    return True if match != None else False

def ask_question(question: str) -> bool:
    '''Asks a yes/no question which is given as input'''
    
    while True:
        answer = input(question + " yes/no ")
        if answer.lower() == "yes":
            return True
        elif answer.lower() == "no":
            return False
        else:
            print("Only yes or no answers accepted, try again")

def collect_user_input(categories: list) -> list:
    '''Creates a list of dictionaries, each dict for one transaction'''

    new_rows = []
    while True:
        row = dict()
        while True:
            row["date"] = input("\nWhat's the date of this purchase? Please use the format dd/mm/yyyy ")
            if validate_date(row["date"]):
                break
            else:
                print("Wrong format, try again")

        while True:
            amount = input("\nWhat's the amount? for example 29.99 or 20 ")
            if validate_price(amount):
                row["amount"] = float(amount)
                break
            else:
                print("Please give the number, decimals are optional, two are allowed, after a dot, eg. 29.99 ")
        
        while True:
            for count, value in enumerate(categories, start=1):
                print(count, value)
            index = int(input("\nChoose the category by writing its number "))
            if 1 <= int(index) <= len(categories):
                row["category"] = categories[index-1]
                break
            else:
                print("Please choose one of the available numbers ")
        
        row["description"] = input("\nAdd a short description for this purchase ")
        new_rows.append(row)
        
        while True:
            another_one = input("Done, would you like to add another purchase? yes/no ")
            if another_one.lower() in ["yes", "no"]:
                break
            else:
                print("Something went wrong, answer yes or no")
        if another_one == "no":
            print(f"cool, the following {len(new_rows)} row(s) will be saved:")
            for row in new_rows:
                print(row)
            return new_rows


def create_new_excel(file_path: str):
    '''Creates a new Excel file with 4 columns,
    saves it to user's desktop'''

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Amount"
    sheet["C1"] = "Category"
    sheet["D1"] = "Description"
    for cell in ["A1", "B1", "C1", "D1"]:
        sheet[cell].fill = PatternFill(start_color="00FF00", patternType="solid")
    workbook.save(filename=file_path)
    workbook.close()

def save_new_rows_to_excel(rows: list, file_path: str):
    '''Accepts a list of transactions the user wants to add
    and inserts them into Excel'''

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    current_index = sheet.max_row + 1
    for row in rows:
        sheet["A"+str(current_index)] = datetime.strptime(row['date'], "%d/%m/%Y").date()
        sheet["B"+str(current_index)] = row['amount']
        sheet["C"+str(current_index)] = row['category']
        sheet["D"+str(current_index)] = row['description']
        current_index +=1
    
    workbook.save(filename=file_path)

def get_stats_options(file_path: str) -> list:
    '''Loops through all excel rows, 
    returns a list of available months which user can get stats for'''

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    options = []
    for row in sheet.iter_rows(min_row=2,min_col=1, max_col=4, values_only=True):
        month_year = datetime.strftime(row[0], "%m/%y")
        if month_year not in options:
            options.append(month_year) 
    # Sort chronologically in case the user did not feed excel with dates in chronological order 
    options.sort(key= lambda date: datetime.strptime(date, "%m/%y"))
    return options

def get_user_request(stats_options: list) -> list:  
    '''Asks if user wants to see stats.
    If no, returns empty list. If yes, prints the available options.
    Returns a sorted list of chosen options'''
    
    while True:
        answer = input("Would you like to see statistics? yes/no ")
        if answer.lower() == "no":
            return []
        elif answer.lower() == "yes":
            break
        else:
            print("Something went wrong, only yes or no answers are valid, try again")
    
    while True:
        print("\nYou can get summary for the following months:")
        for option in stats_options:
            print(option)
        
        user_choices = []
        while True:
            options_chosen = input(
                "\nWrite which months in format mm/YY, seperate them by one space, for example:"
                "\n'01/20 02/20 03/20 04/20' Your choices: "
            )
            
            is_correct = True
            for option in options_chosen.split():
                if option in stats_options:
                    user_choices.append(option)
                else:
                    is_correct = False
                    print(
                        "\n" + option + " is not available."
                        "\nMake sure the option is valid and follows the format mm/yy."
                    )
            if is_correct:
                break
        while True:
            print("\nYou have requested statistics for these months:")
            for option in user_choices:
                print(option)
            want_change = input("\nWould you like to change anything? yes/no ")
            if want_change.lower() == "no":
                user_choices.sort(key= lambda date: datetime.strptime(date, "%m/%y"))
                return user_choices
            elif want_change.lower() == "yes":
                break
            else:
                print("Something went wrong, try again: yes/no")

def get_stats_data(categories: list, user_choices: list, file_path: str) -> list:
    '''Returns a list of dictionaries, each dict for month chosen by the user
    Dict contains month's name, total, top 5 transactions and sums for categories'''

    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    stats_data = []
    for month in user_choices:
        new_dict = {}
        new_dict["month"] = month
        new_dict["total"] = 0
        all_transactions = []
        for row in sheet.iter_rows(min_row=2,min_col=1, max_col=4, values_only=True):
            if month == datetime.strftime(row[0], "%m/%y"):
                new_dict["total"] += row[1]
                all_transactions.append(row)
        all_transactions.sort(key=lambda x: x[1], reverse=True)
        new_dict["top_5"] = all_transactions[:5]
        new_dict["categories"] = {}
        for category in categories:
            new_dict["categories"][category] = sum([i[1] for i in all_transactions if i[2] == category])
        stats_data.append(new_dict)
    print(stats_data)
    return stats_data

def prepare_message(stats_data: list) -> str:
    '''Returns a message string with stats
    for months chosen by the user'''
    
    message = ""
    for month in stats_data:
        message += (
            f"\nHere are statistics for {month['month']}:\n"
            f"\nThe total spent: {month['total']}\n"
            "\nThe highest transactions of the month: \n"
        )

        for num, i in enumerate(month['top_5'], start=1):
            message += f"{num}: {datetime.strftime(i[0], '%d/%m/%y')}, amount: {i[1]}, category: {i[2]}, details: {i[3]}\n"
        
        message += "\nTotals of each category:\n"
        for k, v in month['categories'].items():
            message += f"{k}: {v}\n"
    return message

def plot(categories: list, stats_data: list):
    '''One x axis containing months created.
    Then multiple y axis are created - one for totals,
    one for each category'''
    
    x = [i['month'] for i in stats_data]
    y_totals = [i['total'] for i in stats_data]
    plt.plot(x, y_totals, label='Totals')

    for category in categories:
        new_y = []
        for month in stats_data:      
            new_y.append(month['categories'][category])
        plt.plot(x, new_y, label=category)
    
    plt.xlabel('Months')
    plt.ylabel('Money spent in PLN')
    plt.title('Money spending over months')
    plt.legend()
    plt.show()

def main():
    '''First adds new rows to Excel if user wishes so.
    Then gets all available months and display stats
    for the ones chosen by the user (also optional).
    If more than one month chosen, displays a chart'''

    if os.path.isfile(file_path) == False:
        create_new_excel(file_path)
    if ask_question("Would you like to add any rows to Excel?"):
        new_rows = collect_user_input(categories)
        save_new_rows_to_excel(new_rows, file_path)
    stats_options = get_stats_options(file_path)
    if stats_options != []:
        user_choices = get_user_request(stats_options)
        if user_choices != []:
            stats_data = get_stats_data(categories, user_choices, file_path)
            message = prepare_message(stats_data)
            print(message)
            if len(user_choices) > 1:
                plot(categories, stats_data)
    input("\nHit the enter to exit. Thanks!")

if __name__ == "__main__":
    main()
